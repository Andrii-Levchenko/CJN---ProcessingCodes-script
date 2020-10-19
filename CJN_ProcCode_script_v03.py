import glob
import xlrd
import openpyxl
import re
import datetime
import os

'''Task: to check if ProcessingCode in Excel-file matches ProcessingCode in dxf-file (plate contours)'''

today = datetime.datetime.today()

def report_create(report_name):
    '''create report-files with given name'''
    filename=(today.strftime("%Y-%m-%d-%H.%M.%S")+' - '+report_name+'.txt')
    with open(filename, 'w', encoding="utf-8") as file_object:
        return file_object.write('')

def report_update(report_name, part_number_in_dxf,correctness,file):
    '''add the information about plates parts (with adress of the each file) to report-files with given name'''
    filename=(today.strftime("%Y-%m-%d-%H.%M.%S")+' - '+report_name+'.txt')
    with open(filename, 'a', encoding="utf-8") as file_object:
        return file_object.write('Plate part number '+part_number_in_dxf+' has '+correctness+' ProcessingCode in dxf file!\nPath: '+file+'\n\n')

def files_searching(path, extension):
    '''search files with given extension in given folder and its subfolders.
        create a list of found files'''
    files_list = glob.glob(path + '/**/*.'+extension, recursive = True)
    return files_list

def check_element_in_dxf(element):
    '''check the number of given element in dxf file'''
    element_check = re.search(r'\['+element+'\](.*)\n',read_file).group()
    element_number_in_dxf = element_check[3:].rstrip()
    return element_number_in_dxf

def elements_list_xlsx(element):
    '''create list of elements with given name'''
    for i in range (1,sheet.max_column):
        value = sheet.cell(row=1, column=i).value
        if value == element:
            elements_list = []
            for j in range(2, sheet.max_row):
                new_val = str(sheet.cell(row=j, column=i).value)
                elements_list.append(new_val)
    return elements_list

report_create('SuccesReport')
report_create('FailureReport')

#dictionary - encryption table for ProcessingCodes
ProcessingCodes={'1000':'S',
                '1001':'S K',
                '1002':'S L',
                '1003':'S K L',
                '2000':'S V',
                '2001':'S V L',
                '400':'R',
                '402':'R B',
                '407':'R',
                '606':'G',
                '607':'G L',
                '22':'R C',
                '230':'S V L',
                '20':'S L',
                '0':'R',
                '408':'R C',
                '409':'R C'}

section_for_checking = input('Input the section number: ')

path = input('Input folder adress to find the Process-list : ') #we will search in this directory and its subdirectories

try:      #if the report is made in a new Excel
    xl_files_list = files_searching(path,'xlsx')
    if xl_files_list == []: # if the list is empty - i.e. no .xlsx files in specified folder and subfolders
        raise Exception
    else:
        for xlfile in xl_files_list:
            if 'c-job_process' in xlfile:
                wb=openpyxl.load_workbook(xlfile)
                sheet=wb.active
                for i in range (1,sheet.max_column):
                    value = sheet.cell(row=1, column=i).value
                    if value == 'Block':
                        block = str(sheet.cell(row=2, column=i).value)

                if block != section_for_checking: # so that information is not read from excel belonging to other sections
                    continue

                parts_list = elements_list_xlsx('Part') # create a list of part numbers for further comparison with the value in the autocad
                codes_list = elements_list_xlsx('Processingcode')     # create a list of letter codes to translate them according to the encryption table

            else:
                raise Exception

except Exception:  #if the report is made in old Excel
    xl_files_list = files_searching(path,'xls')
    for xlfile in xl_files_list:
        if 'c-job_process' in xlfile:
            wb = xlrd.open_workbook(xlfile,formatting_info=True)
            sheet = wb.sheet_by_index(0)

# variable to check if a part is from this block (if there are drawings of parts of different blocks on the computer or several 'process-list'-files from different blocks
            row=0
            for col in range(sheet.ncols):
                if sheet.cell_value(row, col) == 'Block':
                    Block=str(int(sheet.cell_value(row+1, col)))

            if Block != section_for_checking: # so that information is not read from excel belonging to other sections
                continue

# create a list of part numbers for further comparison with the value in the autocad
            row=0
            for col in range(sheet.ncols):
                if sheet.cell_value(row, col) == 'Part':
                    parts_list = []
                    for row in range(1,sheet.nrows):
                        new_val=str(int(sheet.cell_value(row, col)))
                        parts_list.append(new_val)

# create a list of letter codes to translate them according to the encryption table
            row = 0
            for col in range(sheet.ncols):
                if sheet.cell_value(row, col) == 'Processingcode':
                    codes_list = []
                    for row in range(1,sheet.nrows):
                        new_val=sheet.cell_value(row, col)
                        codes_list.append(new_val)

# function to get a key by value from a dictionary of a code encryption table
def get_key(ProcessingCodes, value):
    for k, v in ProcessingCodes.items():
        if v == value:
            return k

# list of numeric codes after translation by encryption table
num_codes_list=list()
for item in codes_list:
    num_codes_list.append(get_key(ProcessingCodes,item))

# dictionary 'part number - numeric code'
Part_Code_Dictionary = dict(zip(parts_list,num_codes_list))


path=input('Input folder adress to find the DXF-file : ') # we will search for dxf in this directory and its subdirectories
acad_files = files_searching(path,'dxf')
for file in acad_files:
    fhand=open(file)
    read_file = fhand.read()

    if '[S]' and '[P]' and '[c]' not in read_file: # check that there is no further search for unnecessary files
        continue

    else:  # check if the open part is in the block given in 'section_for_checking'
        block_number_in_dxf = check_element_in_dxf('S')

        if block_number_in_dxf != section_for_checking:
            continue

        else:
            part_number_in_dxf = check_element_in_dxf('P') # read the part number

            proc_code_in_dxf = check_element_in_dxf('c')# read the ProcessingCode of the part

            if (Part_Code_Dictionary[part_number_in_dxf]) != proc_code_in_dxf: # compare with the information from the Excel file using the previously created dictionary
                report_update('FailureReport', part_number_in_dxf,'incorrect',file)

            else:
                report_update('SuccesReport', part_number_in_dxf,'correct',file)

reports_adress = os.path.abspath(path)
print ('Reports have been generated. Look for them in the next folder : '+ reports_adress)

close=input('Push ENTER to close programm')
